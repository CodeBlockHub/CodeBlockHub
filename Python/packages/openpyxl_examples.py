from openpyxl import Workbook, load_workbook
'''
Below are some examples demonstrating various operations you can perform using the openpyxl package in Python.
'''

# Example 1: Create a new Excel workbook and add data
wb = Workbook()
ws1 = wb.active
ws1.title = "Sheet1"
ws1['A1'] = 'Hello'
ws1['A2'] = 'World'
wb.save("example1.xlsx")

# Example 2: Load an existing Excel workbook and read data
wb = load_workbook('example1.xlsx')
ws2 = wb.active
print(ws2['A1'].value)  # Output should be 'Hello'
print(ws2['A2'].value)  # Output should be 'World'

# Example 3: Update a cell in an existing Excel workbook
wb = load_workbook('example1.xlsx')
ws3 = wb.active
ws3['A1'] = 'Hi'
wb.save('example1.xlsx')

# Example 4: Add a new worksheet and populate it
wb = load_workbook('example1.xlsx')
ws4 = wb.create_sheet(title="Sheet2")
ws4['A1'] = 'New'
ws4['A2'] = 'Sheet'
wb.save('example1.xlsx')

# Example 5: Add formulas
wb = load_workbook('example1.xlsx')
ws5 = wb.active
ws5['B1'] = 10
ws5['B2'] = 20
ws5['B3'] = '=SUM(B1:B2)'
wb.save('example1.xlsx')

